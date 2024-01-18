import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook, load_workbook 

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

url = "https://www.imdb.com/"
driver.get(url)
time.sleep(2)


find=driver.find_element(By.ID, "imdbHeader-navDrawerOpen")
find.click()
time.sleep(3)
find=driver.find_element(By.XPATH, "//span[@class='ipc-list-item__text' and text()='Most Popular Movies']")
find.click()
time.sleep(2)

movies = []
ratings = []

find=driver.find_element(By.XPATH, "//h3[@class='ipc-title__text' and text()='Saltburn']")
movies.append(find.text)
find=driver.find_element(By.XPATH, '//span[@aria-label="IMDb rating: 7.1"]')
ratings.append(find.text)
find=driver.find_element(By.XPATH, "//h3[@class='ipc-title__text' and text()='Poor Things']")
movies.append(find.text)
find=driver.find_element(By.XPATH, '//span[@aria-label="IMDb rating: 8.4"]')
ratings.append(find.text)
find=driver.find_element(By.XPATH, "//h3[@class='ipc-title__text' and text()='Society of the Snow']")
movies.append(find.text)
find=driver.find_element(By.XPATH, '//span[@aria-label="IMDb rating: 7.9"]')
ratings.append(find.text)
find=driver.find_element(By.XPATH, "//h3[@class='ipc-title__text' and text()='Mean Girls']")
movies.append(find.text)
find=driver.find_element(By.XPATH, '//span[@aria-label="IMDb rating: 6.4"]')
ratings.append(find.text)
find=driver.find_element(By.XPATH, "//h3[@class='ipc-title__text' and text()='12th Fail']")
movies.append(find.text)
find=driver.find_element(By.XPATH, '//span[@aria-label="IMDb rating: 9.2"]')
ratings.append(find.text)
find=driver.find_element(By.XPATH, "//h3[@class='ipc-title__text' and text()='Oppenheimer']")
movies.append(find.text)
find=driver.find_element(By.XPATH, '//span[@aria-label="IMDb rating: 8.4"]')
ratings.append(find.text)
find=driver.find_element(By.XPATH, "//h3[@class='ipc-title__text' and text()='The Holdovers']")
movies.append(find.text)
find=driver.find_element(By.XPATH, '//span[@aria-label="IMDb rating: 8.0"]')
ratings.append(find.text)
find=driver.find_element(By.XPATH, "//h3[@class='ipc-title__text' and text()='Anyone But You']")
movies.append(find.text)
find=driver.find_element(By.XPATH, '//span[@aria-label="IMDb rating: 6.6"]')
ratings.append(find.text)
find=driver.find_element(By.XPATH, "//h3[@class='ipc-title__text' and text()='Killers of the Flower Moon']")
movies.append(find.text)
find=driver.find_element(By.XPATH, '//span[@aria-label="IMDb rating: 7.7"]')
ratings.append(find.text)
find=driver.find_element(By.XPATH, "//h3[@class='ipc-title__text' and text()='The Beekeeper']")
movies.append(find.text)
find=driver.find_element(By.XPATH, '//span[@aria-label="IMDb rating: 6.8"]')
ratings.append(find.text)
time.sleep(5)

find=driver.find_element(By.ID, "imdbHeader-navDrawerOpen")
find.click()
time.sleep(3)
find=driver.find_element(By.XPATH, "//span[@class='ipc-list-item__text' and text()='Most Popular TV Shows']")
find.click()
time.sleep(2)

shows = []
showratings = []

find=driver.find_element(By.XPATH, "//h3[@class='ipc-title__text' and text()='Fool Me Once']")
shows.append(find.text)
find=driver.find_element(By.XPATH, '//span[@aria-label="IMDb rating: 6.9"]')
showratings.append(find.text)
find=driver.find_element(By.XPATH, "//h3[@class='ipc-title__text' and text()='Echo']")
shows.append(find.text)
find=driver.find_element(By.XPATH, '//span[@aria-label="IMDb rating: 6.1"]')
showratings.append(find.text)
find=driver.find_element(By.XPATH, "//h3[@class='ipc-title__text' and text()='Reacher']")
shows.append(find.text)
find=driver.find_element(By.XPATH, '//span[@aria-label="IMDb rating: 8.1"]')
showratings.append(find.text)
find=driver.find_element(By.XPATH, "//h3[@class='ipc-title__text' and text()='Fargo']")
shows.append(find.text)
find=driver.find_element(By.XPATH, '//span[@aria-label="IMDb rating: 8.9"]')
showratings.append(find.text)
find=driver.find_element(By.XPATH, "//h3[@class='ipc-title__text' and text()='True Detective']")
shows.append(find.text)
find=driver.find_element(By.XPATH, '//span[@aria-label="IMDb rating: 8.9"]')
showratings.append(find.text)
find=driver.find_element(By.XPATH, "//h3[@class='ipc-title__text' and text()='The Brothers Sun']")
shows.append(find.text)
find=driver.find_element(By.XPATH, '//span[@aria-label="IMDb rating: 7.7"]')
showratings.append(find.text)
find=driver.find_element(By.XPATH, "//h3[@class='ipc-title__text' and text()='Percy Jackson and the Olympians']")
shows.append(find.text)
find=driver.find_element(By.XPATH, '//span[@aria-label="IMDb rating: 7.4"]')
showratings.append(find.text)
find=driver.find_element(By.XPATH, "//h3[@class='ipc-title__text' and text()='The Bear']")
shows.append(find.text)
find=driver.find_element(By.XPATH, '//span[@aria-label="IMDb rating: 8.6"]')
showratings.append(find.text)
find=driver.find_element(By.XPATH, "//h3[@class='ipc-title__text' and text()='Succession']")
shows.append(find.text)
find=driver.find_element(By.XPATH, '//span[@aria-label="IMDb rating: 8.9"]')
showratings.append(find.text)
find=driver.find_element(By.XPATH, "//h3[@class='ipc-title__text' and text()='Berlin']")
shows.append(find.text)
find=driver.find_element(By.XPATH, '//span[@aria-label="IMDb rating: 7.1"]')
showratings.append(find.text)

wb=load_workbook('FILMAS.xlsx')
ws=wb.active

for i in range(len(movies)):
    ws['B'+str(i+4)]=movies[i]
    ws['E'+str(i+4)]=ratings[i]

for i in range(len(movies)):
    ws['H'+str(i+4)]=shows[i]
    ws['K'+str(i+4)]=showratings[i]

wb.save('FILMAS.xlsx')

input()